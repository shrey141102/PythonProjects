import pygame
import random
import tkinter as tk
from tkinter import simpledialog
from openpyxl import load_workbook, Workbook
import os

# Constants
WIDTH, HEIGHT = 800, 600
MOVE_INCREMENT = 20
FPS = 15

# Colors
BLACK = (0, 0, 0)
GREEN = (0, 255, 0)
RED = (255, 0, 0)
WHITE = (255, 255, 255)

# Initialize Pygame
pygame.init()
win = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption("Snake Game")
clock = pygame.time.Clock()

# Snake initial position and speed
snake = [(100, 100), (80, 100), (60, 100)]
direction = "Right"
food_position = (200, 200)
score = 0
highest_score = 0
player_name = ""

def generate_food():
    return random.randrange(0, WIDTH - MOVE_INCREMENT, MOVE_INCREMENT), random.randrange(0, HEIGHT - MOVE_INCREMENT, MOVE_INCREMENT)

def draw_snake(snake):
    for segment in snake:
        pygame.draw.rect(win, GREEN, (segment[0], segment[1], MOVE_INCREMENT, MOVE_INCREMENT))

def draw_food(food_position):
    pygame.draw.rect(win, RED, (food_position[0], food_position[1], MOVE_INCREMENT, MOVE_INCREMENT))

def move():
    global direction, score, food_position, highest_score
    head_x, head_y = snake[0]

    if direction == "Left":
        head_x -= MOVE_INCREMENT
    elif direction == "Right":
        head_x += MOVE_INCREMENT
    elif direction == "Up":
        head_y -= MOVE_INCREMENT
    elif direction == "Down":
        head_y += MOVE_INCREMENT

    snake.insert(0, (head_x, head_y))

    # Check collision with food
    if (head_x, head_y) == food_position:
        score += 10
        food_position = generate_food()
    else:
        snake.pop()

    # Check collision with walls or itself
    if (
        head_x < 0
        or head_x >= WIDTH
        or head_y < 0
        or head_y >= HEIGHT
        or any(segment == (head_x, head_y) for segment in snake[1:])
    ):
        if score > highest_score:
            highest_score = score
            save_high_score(player_name, highest_score)  # Save highest score to Excel
        snake.clear()
        snake.append((100, 100))
        direction = "Right"
        food_position = generate_food()  # Reset food position
        score = 0

def save_high_score(name, score):
    file_exists = os.path.exists("high_scores.xlsx")
    wb = load_workbook("high_scores.xlsx") if file_exists else Workbook()
    sheet = wb.active if file_exists else wb.create_sheet("Sheet")
    
    if name not in [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]:
        sheet.append([name])
    
    name_row = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)].index(name) + 1
    score_col = sheet.max_column + 1
    
    sheet.cell(row=name_row, column=score_col, value=score)
    wb.save("high_scores.xlsx")

def main():
    global direction, player_name, highest_score
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    player_name = simpledialog.askstring("Player Name", "Enter your name:")
    if player_name:
        player_name = player_name.lower()  # Convert player name to lowercase
    else:
        player_name = "unknown"  # Set default name if user cancels input

    run = True
    while run:
        clock.tick(FPS)
        win.fill(BLACK)  # Set background color to black

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                run = False
            elif event.type == pygame.KEYDOWN:
                if event.key == pygame.K_LEFT and direction != "Right":
                    direction = "Left"
                elif event.key == pygame.K_RIGHT and direction != "Left":
                    direction = "Right"
                elif event.key == pygame.K_UP and direction != "Down":
                    direction = "Up"
                elif event.key == pygame.K_DOWN and direction != "Up":
                    direction = "Down"

        move()
        draw_snake(snake)
        draw_food(food_position)

        font = pygame.font.SysFont(None, 36)
        text = font.render(f'Highest Score: {highest_score}', True, WHITE)
        win.blit(text, (10, 10))

        pygame.display.update()

    pygame.quit()

if __name__ == "__main__":
    main()
